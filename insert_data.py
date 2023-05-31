# Feeds csv data into Excel file for graph generation
#
# Copyright 2022-2023 Joaquin M Lopez Munoz.
# Distributed under the Boost Software License, Version 1.0.
# (See accompanying file LICENSE_1_0.txt or copy at
# http://www.boost.org/LICENSE_1_0.txt)
#
# See https://github.com/joaquintides/boost_unordered_benchmark
# for project home page.

import argparse
import io
import openpyxl
import re
import sys

parser=argparse.ArgumentParser(
    description="Feeds csv data into Excel file for graph generation")
parser.add_argument("input_file",help="csv file")
parser.add_argument("section",help="section header in the csv file")
parser.add_argument("excel_file",help="Excel file")
parser.add_argument("sheet_name",help="Sheet name")
parser.add_argument("-e",dest="encoding",default=None,
  help="encoding (default none)")
args=parser.parse_args()

with io.open(args.input_file,"r",encoding=args.encoding) as filein:
  lines=filein.readlines()
  section_found=False
  for lines_read,line in enumerate(lines,1):
    if line.rstrip()==args.section+":":
      lines_read+=1 # skip header
      section_found=True
      break
  if not section_found: sys.exit("section \"{}\" not found".format(args.section))

  wb=openpyxl.load_workbook(args.excel_file)
  ws=wb[args.sheet_name]
  for row in range(2,ws.max_row+1):
    ws.cell(row=row,column=1).value=None
    ws.cell(row=row,column=2).value=None
    ws.cell(row=row,column=3).value=None
    ws.cell(row=row,column=4).value=None

  pattern=re.compile(r"([0-9]+);([0-9.]+);([0-9.]+);([0-9.]+)")
  for row,line in enumerate(lines[lines_read:],2):
    m=pattern.match(line)
    if not m: break
    ws.cell(row=row,column=1).value=int(m.group(1))
    ws.cell(row=row,column=2).value=float(m.group(2))
    ws.cell(row=row,column=3).value=float(m.group(3))
    ws.cell(row=row,column=4).value=float(m.group(4))
    
  wb.save(args.excel_file)
